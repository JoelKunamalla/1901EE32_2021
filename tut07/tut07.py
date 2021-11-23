from collections import defaultdict
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import csv


def feedback_not_submitted():

	with open("course_master_dont_open_in_excel.csv" , "r") as f:
	    d = csv.reader(f)
	    l1={}
	    for x in d:
	        if x[0]=="subno":
	            continue
	        for v in x:
	            fii= x[2].split("-")
	            l1[x[0]]=fii
	
	with open("studentinfo.csv", "r") as f:
	    qw= csv.reader(f)
	    ZX=defaultdict(list)
	    for xy in qw:
	        if xy[0]=="Name":
	                continue
	        ZX[xy[1]].append([ xy[0] ,  xy[8] , xy[9] , xy[10]])

	joel=[]
	l=0
	with open("course_feedback_submitted_by_students.csv" , "r") as f:
	            g1 = csv.reader(f)
	            ss= defaultdict(list)
	            for x1 in g1:
	                l+=1
	                if x1[1]=="stud_email":
	                    continue
	                z=x1[3] +x1[4] +str(x1[5])
	                ss[x1[3]].append(z)
	with open("course_registered_by_all_students.csv" , "r") as f:
	    g= csv.reader(f)
	    l2={}
	    for x in g:
	        if x[0]=="rollno":
	            continue
	        mk= ss[x[0]]
	        zz= l1[x[3]]
	        for i in range(0,3):
	                    if int(zz[i])  > 0:
	                        if x[0] in ZX :
	                            a=ZX[x[0]][0]
	                        else:
	                            a=["NA_IN_STUDENTINFO" , "NA_IN_STUDENTINFO" , "NA_IN_STUDENTINFO", "NA_IN_STUDENTINFO"]
	                        klj=True
	                        for xx in range(0 , len(mk)):
	                            if x[0] + x[3] + str(i+1) == mk[xx]:
	                                klj=False
	                                break
	                        if   klj:
	                                joel.append([x[0] , x[1] , x[2] , x[3]  ,a[0] , a[1] , a[2] ,a[3]])
	if os.path.exists("course_feedback_remaining.xlsx" ):
	    os.remove("course_feedback_remaining.xlsx")
	wb = Workbook()
	sheet = wb.active
	sheet.title = "sheet1"
	wb.save(f"course_feedback_remaining.xlsx")
	wb=load_workbook("course_feedback_remaining.xlsx")
	sheet = wb.active
	sheet.append(["rollno" , "registered_sem" , "scheduled_sem" , "subno" , "Name" ,"email" , "aemail" , "contact"])
	while(len(joel) >0):
	    sheet.append(joel[0])
	    joel.pop(0)
	wb.save(f"course_feedback_remaining.xlsx")
feedback_not_submitted()