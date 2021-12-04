
from types import MethodDescriptorType
from flask import Flask, render_template, request , flash , Markup , send_file
import flask
import os
from werkzeug.utils import secure_filename
from openpyxl.utils import cell
import csv
from openpyxl import Workbook, workbook
from openpyxl import load_workbook
import os
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import shutil
from flask_mail import Mail, Message

app = Flask(__name__)
mail=Mail(app)

@app.route('/')
def get():
    # return 'hello  world!'
    flash("ALERTS will be displayed here")
    return render_template('project1.html')


@app.route('/joel11',  methods=['POST', 'GET'])
def generate_marksheet11():

    if(os.path.exists("./storage")):
            shutil.rmtree("./storage")
    else:
            pass
    os.mkdir("./storage")
    if(os.path.exists("./marksheets")):
            shutil.rmtree("./marksheets")
    else:
            pass
    os.mkdir("./marksheets")
    if flask.request.method == 'POST':
        g = request.files["master_roll"]
        f1 = request.files["response"]
        g.save(os.path.join("storage" , (g.filename)))
        f1.save(os.path.join("storage" , (f1.filename)))
    flash("UPLOADED SUCCESSFULLY ")
    return render_template("project1.html")










@app.route('/joel',  methods=['POST', 'GET'])
def generate_marksheet():
    s= os.listdir("storage")
    print(s)
    if(  os.path.exists("./marksheets")):
       pass     
    else:
        os.mkdir("./marksheets")
    if flask.request.method == 'POST':
        i1 = request.form.get("correct_marks")
        i2 = request.form.get("negative_marks")
    h = float(i1)
    i = float(i2)
    if(i >0):
        flash("please enter valid input")
    else:
    
        f111=0
        
        with open(f"./storage/responses.csv" , "r") as f:
            r= csv.reader(f)
            d=[]
            for x in r:
                if x[6]=="ANSWER" or x[6]=="Roll Number":
                    ANSWER_NAME=x[3]
                    continue
                d.append(x[6])
        with open(f"./storage/responses.csv" , "r") as f:
            r= csv.reader(f)

            for x in r:
                if x[6]=="ANSWER":
                    f111=1
                    break
                
        if f111==1:
            print(11111)
            dicasdf={}
            with open(f"./storage/master_roll.csv" , "r") as f:
                reader=csv.reader(f , delimiter=',')
                d1=[]
                c1=[]
                for x in reader:
                    if x[0]=="roll":
                        pass
                    else:
                        d1.append(x[0])
                        c1.append(x[1])
                        dicasdf[x[0]]=x[1]
                dic={x:[] for x in d}
                dic12={x:[] for x in d}
                dic11={x:[] for x in d}
                excep1={'ANSWER':[]}
                excep12={'ANSWER':[]}
                excep2={'ANSWER':[]}
            mnb=[]
            mnb2=[]
            print(c1)
            print(d)

            with open(f"./storage/responses.csv" , "r") as f:
                reader= csv.reader(f , delimiter=",")
                for x in reader:
                    globbal=len(x)-7

                    correct , wrong , not_attempt=0,0 ,0
                    if x[3]=='Name':
                        pass
                    elif x[6]=="ANSWER":
                        c=x.copy()
                        if(len(x) > 32):

                            for y in range(7 , 32):
                                excep1['ANSWER'].append([x[y] , x[y]])
                            excep2['ANSWER'].append([len(x)-7, 0, 0 , len(x)-7])

                            for y in range(32 , len(x)):
                                excep12['ANSWER'].append([x[y] , x[y]])
                        else:
                            for y in range(7 , len(x)):
                                excep1['ANSWER'].append([x[y] , x[y]])
                            excep2['ANSWER'].append([len(x)-7, 0, 0 , len(x)-7])


                    else:
                        mnb.append(x[6])
                        mnb2.append(x[3])
                        if(len(x) <= 32):

                            for y in range(7 , len(x)):
                                if x[y]==c[y]:
                                    correct+=1
                                elif x[y]=="":
                                    not_attempt+=1
                                else :
                                    wrong+=1
                                dic[x[6]].append([x[y] , c[y]])
                            dic11[x[6]].append([correct , wrong , not_attempt , len(x)-7])
                        else:
                            for y in range(7 , len(x)):
                                if x[y]==c[y]:
                                    correct+=1
                                elif x[y]=="":
                                    not_attempt+=1
                                else :
                                    wrong+=1
                            dic11[x[6]].append([correct , wrong , not_attempt , len(x)-7])
                            for y in range(7 , 32):
                                dic[x[6]].append([x[y] , c[y]])
                            for y in range(32, len(x)):
                                dic12[x[6]].append([x[y] , c[y]])

            clone=d.copy()
            if(os.path.exists("./marksheets")):
                shutil.rmtree("./marksheets")
            else:
                pass
            os.mkdir("./marksheets")


            absentiess=[]
            absentiess2=[]
            for xoy in d1:
                    if xoy in mnb:
                        continue
                    else:
                        absentiess.append(xoy)
                        absentiess2.append(dicasdf[xoy])
            while (len(d) >0):
                if( not os.path.exists(f"./marksheets/{d[0]}.xlsx")):
                    wb = Workbook()
                    sheet= wb.active
                    sheet.title="quiz"
                    wb.save(f"./marksheets/{d[0]}.xlsx")
                    d.pop(0)
            # dic11.pop('ANSWER')
            checking=0
            while(len(clone)>0):
                wb= load_workbook(f"./marksheets/{clone[0]}.xlsx")
                sheet=wb.active
                img =Image('./templates/Picture.png')
                img.anchor="A1"
                img.width = 627.5
                img.height = 81
                sheet.add_image(img)
                sheet.merge_cells('A5:E5')
                sheet['A5']="Mark Sheet"
                currentCell = sheet['A5']
                currentCell.alignment = Alignment(horizontal='center')
                currentCell.font=Font(name="century" , size=18 , underline="single")




                if checking==0:
                    sheet.append(["Name:" , ANSWER_NAME , None , "Exam:" , "quiz"] )
                    sheet.append(["Roll Number:" ,"ANSWER"])
                    sheet.append([None])
                    sheet.append([None, "Right" , "Wrong" , "Not Attempt" , "Max"])
                    sheet.append(["NO." ,excep2["ANSWER"][0][0] , excep2["ANSWER"][0][1] ,   excep2["ANSWER"][0][2] ,   excep2["ANSWER"][0][3]   ])
                    sheet.append(["Marking" ,h , i ,   excep2["ANSWER"][0][2]  ])
                    sheet.append(["Total" ,h*excep2["ANSWER"][0][0] , i*excep2["ANSWER"][0][1] , None,  f"{h*excep2['ANSWER'][0][0] + i*excep2['ANSWER'][0][1]}/{excep2['ANSWER'][0][3]*h}" ])
                    sheet.append([None])
                    sheet.append([None])
                    sheet.append(["Student Ans" , "Correct Ans" , None , "Student Ans" , "Correct Ans"])
                    while(len(excep1["ANSWER"]) > 0 or len(excep12["ANSWER"] )>0):
                        if(len(excep12["ANSWER"] )>0):
                            sheet.append([excep1["ANSWER"][0][0] , excep1["ANSWER"][0][1]     , None, excep12["ANSWER"][0][0] , excep12["ANSWER"][0][1] ])
                            excep1["ANSWER"].pop(0)
                            excep12["ANSWER"].pop(0)
                        else:
                            sheet.append([excep1["ANSWER"][0][0] , excep1["ANSWER"][0][1]   ])
                            excep1["ANSWER"].pop(0)

                else:
                    checking=1
                    sheet.append(["Name:" , dicasdf[clone[0]] , None , "Exam:" , "quiz"] )
                    sheet.append(["Roll Number:" ,clone[0]])
                    sheet.append([None])
                    sheet.append([None, "Right" , "Wrong" , "Not Attempt" , "Max"])
                    print(dic11[clone[0]] , clone[0])
                    print(1)
                    sheet.append(["NO." ,dic11[clone[0]][0][0] , dic11[clone[0]][0][1] , dic11[clone[0]][0][2] , dic11[clone[0]][0][3]   ])
                    sheet.append(["Marking" ,h , i , 0 ])
                    sheet.append(["Total" ,h*dic11[clone[0]][0][0] , i*dic11[clone[0]][0][1] , None,  f"{h*dic11[clone[0]][0][0] + i*dic11[clone[0]][0][1]}/{dic11[clone[0]][0][3]*h}" ])
                    sheet.append([None])
                    sheet.append([None])
                    sheet.append(["Student Ans" , "Correct Ans" , None , "Student Ans" , "Correct Ans"])
                    while(len(dic[clone[0]]) > 0  or len(dic12[clone[0]]) >0):
                        if(len(dic12[clone[0]]) >0):

                            sheet.append([dic[clone[0]][0][0] , dic[clone[0]][0][1]     , None,dic12[clone[0]][0][0] , dic12[clone[0]][0][1]  ])
                            dic[clone[0]].pop(0)
                            dic12[clone[0]].pop(0)
                        else:
                            sheet.append([dic[clone[0]][0][0] , dic[clone[0]][0][1] ])
                            dic[clone[0]].pop(0)


                sheet.column_dimensions['A'].width = 20
                sheet.column_dimensions['B'].width = 20
                sheet.column_dimensions['C'].width = 20
                sheet.column_dimensions['D'].width = 20
                sheet.column_dimensions['E'].width=20



                for r in sheet.iter_rows(min_row=6 , max_row=7):
                    for x in r:
                        if x.value =="Name:" or x.value =="Exam:" or x.value =="Roll Number:" :
                            x.alignment = Alignment(horizontal='right')
                            x.font=Font(name="century" , size=12)
                        elif x.value== None:
                            pass
                        elif x.row ==6 or x.row==7:
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True)
                for r in sheet.iter_rows(min_row=9 , max_row=15 , min_col=1 ,max_col=5):
                    for x in r:
                        if type(x.value)== str:
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True)
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                        elif( type(x.value)==int or  type(x.value)==float ) and x.column==2:
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True , color="00339966")
                        elif (type(x.value)==int  or type(x.value)==float) and x.column==3:
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True , color="00FF0000")
                        elif x.row==13 or x.row==14:
                            pass
                        else:
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True)
                for r in sheet.iter_rows(min_row=16 , max_row=sheet.max_row , min_col=2 ,max_col=2):
                    for x in r:
                        thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                        x.border=thin_border
                        x.alignment = Alignment(horizontal='center')
                        x.font=Font(name="century" , size=12 , bold=True , color="0000FF" )
                for r in sheet.iter_rows(min_row=16 , max_row=sheet.max_row , min_col=5 ,max_col=5):
                    for x in r:
                        thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                        x.border=thin_border
                        x.alignment = Alignment(horizontal='center')
                        x.font=Font(name="century" , size=12 , bold=True , color="0000FF" )
                for r in sheet.iter_rows(min_row=16 , max_row=sheet.max_row , min_col=1 ,max_col=1):
                    for x in r:
                        if x.value == sheet.cell(row=x.row , column=2).value:
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True , color="00339966")
                        else:
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True , color="00FF0000")
                for r in sheet.iter_rows(min_row=16 , max_row=sheet.max_row , min_col=4 ,max_col=4):
                    for x in r:
                        if x.value == sheet.cell(row=x.row , column=5).value:
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True , color="00339966")
                        else:
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True , color="00FF0000")




                c11=sheet["E12"]
                thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                c11.border=thin_border
                c11.alignment = Alignment(horizontal='center')
                c11.font=Font(name="century" , size=12 , bold=True , color="0000FF" )

                for r in sheet.iter_rows(min_row=16 , max_row=sheet.max_row , min_col=5 ,max_col=5):
                    for x in r:
                        if x.value == None:
                            x.border = Border(None)
                            c11=sheet.cell(row= x.row , column=4)
                            c11.border=Border(None)

                d1=sheet["C15"]
                d1.border=Border(None)


                sheet.insert_cols(7)
                if checking==0:
                    wb.save(f"./marksheets/{'ANSWER'}.xlsx")
                else:

                    wb.save(f"./marksheets/{clone[0]}.xlsx")
                if checking==1:
                    clone.pop(0)
                checking=1
                c1.pop(0)



            absentiess.pop(0)
            absentiess2.pop(0)
            clone11= absentiess.copy()
            print("11")
            while (len(absentiess) >0):
                if( not os.path.exists(f"./marksheets/{absentiess[0]}.xlsx")):
                    wb = Workbook()
                    sheet= wb.active
                    sheet.title="quiz"
                    wb.save(f"./marksheets/{absentiess[0]}.xlsx")
                    absentiess.pop(0)


            print("12")

            clone111=clone11.copy()
            print(clone111)
            while len(clone111)>0:
                if(  os.path.exists(f"./marksheets/{clone111[0]}.xlsx")):
                    wb = Workbook()
                    sheet= wb.active
                    sheet.title="quiz"
                    wb.save(f"./marksheets/{clone111[0]}.xlsx")
                    clone111.pop(0)



            print("13")
            while(len(clone11)>0 and len(absentiess2)>0):
                wb= load_workbook(f"./marksheets/{clone11[0]}.xlsx")
                sheet=wb.active
                img =Image('./templates/Picture.png')
                img.anchor="A1"
                img.width = 627.5
                img.height = 81
                sheet.add_image(img)
                sheet.merge_cells('A5:E5')
                sheet['A5']="Mark Sheet"
                currentCell = sheet['A5']
                currentCell.alignment = Alignment(horizontal='center')
                currentCell.font=Font(name="century" , size=18 , underline="single")

                sheet.append(["Name:" , absentiess2[0] , None , "Exam:" , "quiz"] )
                sheet.append(["Roll Number:" ,clone11[0]])
                sheet.append([None])
                sheet.append([None, "Right" , "Wrong" , "Not Attempt" , "Max"])
                sheet.append(["NO." ,0 , 0 ,  globbal ,   globbal   ])
                sheet.append(["Marking" ,h , i , 0  ])
                sheet.append(["Total" ,0 , 0, None, "ABSENT"  ])
                sheet.append([None])




                print("15")
                sheet.column_dimensions['A'].width = 20
                sheet.column_dimensions['B'].width = 20
                sheet.column_dimensions['C'].width = 20
                sheet.column_dimensions['D'].width = 20
                sheet.column_dimensions['E'].width=20
                for r in sheet.iter_rows(min_row=6 , max_row=7):
                    for x in r:
                        if x.value =="Name:" or x.value =="Exam:" or x.value =="Roll Number:" :
                            x.alignment = Alignment(horizontal='right')
                            x.font=Font(name="century" , size=12)
                        elif x.value== None:
                            pass
                        elif x.row ==6 or x.row==7:
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True)
                for r in sheet.iter_rows(min_row=9 , max_row=15 , min_col=1 ,max_col=5):
                    for x in r:
                        if type(x.value)== str:
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True)
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                        elif type(x.value)==int and x.column==2:
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True , color="00339966")
                        elif type(x.value)==int and x.column==3:
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True , color="00FF0000")
                        elif x.row==13 or x.row==14:
                            pass
                        else:
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True)
                for r in sheet.iter_rows(min_row=16 , max_row=sheet.max_row , min_col=2 ,max_col=2):
                    for x in r:
                        thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                        x.border=thin_border
                        x.alignment = Alignment(horizontal='center')
                        x.font=Font(name="century" , size=12 , bold=True , color="0000FF" )
                for r in sheet.iter_rows(min_row=16 , max_row=sheet.max_row , min_col=5 ,max_col=5):
                    for x in r:
                        thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                        x.border=thin_border
                        x.alignment = Alignment(horizontal='center')
                        x.font=Font(name="century" , size=12 , bold=True , color="0000FF" )
                for r in sheet.iter_rows(min_row=16 , max_row=sheet.max_row , min_col=1 ,max_col=1):
                    for x in r:
                        if x.value == sheet.cell(row=x.row , column=2).value:
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True , color="00339966")
                        else:
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True , color="00FF0000")
                for r in sheet.iter_rows(min_row=16 , max_row=sheet.max_row , min_col=4 ,max_col=4):
                    for x in r:
                        if x.value == sheet.cell(row=x.row , column=5).value:
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True , color="00339966")
                        else:
                            thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                            x.border=thin_border
                            x.alignment = Alignment(horizontal='center')
                            x.font=Font(name="century" , size=12 , bold=True , color="00FF0000")
                c11=sheet["E12"]
                thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'),  top=Side(style='thin'),  bottom=Side(style='thin'))
                c11.border=thin_border
                c11.alignment = Alignment(horizontal='center')
                c11.font=Font(name="century" , size=12 , bold=True , color="0000FF" )
                for r in sheet.iter_rows(min_row=16 , max_row=sheet.max_row , min_col=5 ,max_col=5):
                    for x in r:
                        if x.value == None:
                            x.border = Border(None)
                            c11=sheet.cell(row= x.row , column=4)
                            c11.border=Border(None)
                d1=sheet["C15"]
                d1.border=Border(None)



                wb.save(f"./marksheets/{clone11[0]}.xlsx")
                clone11.pop(0)
                absentiess2.pop(0)

            print("14")


            flash("generated rollwise marksheet")
        else:
            flash("NO ANSWER ROW")
    return render_template("project1.html")






print("233")
@app.route("/jol" , methods=["POST" , "GET"])
def downloadd():
    print("236")
    if flask.request.method=='POST':
        shutil.make_archive("marksZIP", 'zip',"marksheets" )
    flash("downloaded successfully!")
    return send_file("marksZIP.zip", as_attachment=True, mimetype='zip')



@app.route('/kjl', methods=["POST", "GET"])
def consise_marksheet():
    if flask.request.method == 'POST':
       
        i1 = request.form.get("correct_marks")
        i2 = request.form.get("negative_marks")
    h = float(i1)
    i = float(i2)
    if(os.path.exists("./marksheets")):
        pass
    else:
        os.mkdir("./marksheets")

    if(not os.path.exists("./marksheets/consise_marksheet.xlsx")):
        pass
    else:
        os.remove("./marksheets/consise_marksheet.xlsx")
    wb = Workbook()
    sheet = wb.active
    sheet.title = "consise_marksheet"
    wb.save(f"./marksheets/consise_marksheet.xlsx")



    with open(f"./storage/master_roll.csv" , "r") as f:
                reader=csv.reader(f , delimiter=',')
                d1=[]
                c1=[]
                disc={}
                for x in reader:
                    if x[0]=="roll":
                        pass
                    else:
                        d1.append(x[0])
                        disc[x[0]]=x[1]


    mob=[]
    with open(f"./storage/responses.csv" , "r") as f:
                reader= csv.reader(f , delimiter=",")
                for x in reader:

                    correct , wrong , not_attempt=0,0 ,0
                    if x[3]=='Name':
                        pass
                    elif x[6]=="ANSWER":
                       pass
                    else:
                        mob.append(x[6])



    absentie=[]
    absentie2=[]
    for xoy in d1:
        if xoy in mob:
                        continue
        else:
                        absentie.append(xoy)
                        absentie2.append(disc[xoy])

    



    with open(f"./storage/responses.csv", "r") as f:
        reader = csv.reader(f, delimiter=',')
        d = []
        c1 = []
        for x in reader:
            if x[3] == "Name" :
                pass
            else:
                d.append(x[6])
                c1.append(x[1])

    cdic = {x: [] for x in d}
    cdic1 = {x: [] for x in d}
    first = 0
    # wb = load_workbook(f"./marksheets/{d[1]}.xlsx")
    # sheet = wb.active
    # h = int(sheet.cell(row=11, column=2).value)
    # i = int(sheet.cell(row=11, column=3).value)
    # wb.save(f"./marksheets/{d[1]}.xlsx")
    with open("./storage/responses.csv") as f:
        reader = csv.reader(f, delimiter=",")
        for x in reader:
            correct, wrong, not_attempt = 0, 0, 0
            if x[3] == 'Name':
                first += len(x)-7
                pass
            elif x[6] == "ANSWER":
                c = x.copy()
                for y in range(7, len(x)):
                    cdic['ANSWER'].append([x[y], x[y]])
                cdic1['ANSWER'].append([len(x)-7, 0, 0])

            else:
                for y in range(7, len(x)):
                    if x[y] == c[y]:
                        correct += 1
                    elif x[y] == "":
                        not_attempt += 1
                    else:
                        wrong += 1
                    cdic[x[6]].append([x[y], c[y]])
                cdic1[x[6]].append([correct, wrong, not_attempt])
    wb = load_workbook("./marksheets/consise_marksheet.xlsx")
    sheet = wb.active
    with open(f"./storage/responses.csv", "r") as f:
        reader = csv.reader(f, delimiter=",")
        for x in reader:
            sheet.append(x)

    f11 = d.copy()
    sheet.insert_cols(7)
    print(  d[0] , cdic1)
    for cell in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=6, max_col=6):
        for x in cell:
           
            if(x.row == 1):
                sheet.cell(row=x.row, column=7).value = "Scores_after_negative"
            else:
                print(d[0])
                print(d[0] , cdic1['ANSWER'])
                sheet.cell(
                    row=x.row, column=7).value = f"{h*cdic1[d[0]][0][0] + i*cdic1[d[0]][0][1]}/{first*h}"
                d.pop(0)
    for cell in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=37, max_col=37):
        for x in cell:
            if(x.row == 1):
                sheet.cell(row=x.row, column=37).value = "Status_Ans"
            else:
                sheet.cell(
                    row=x.row, column=37).value = f"[{cdic1[f11[0]][0][0]},{cdic1[f11[0]][0][1]},{cdic1[f11[0]][0][2]}]"
                f11.pop(0)

    

    while(len(absentie) >0 and  len(absentie2) >0):
        sheet.append(["ABSENT" ,"ABSENT" ,"ABSENT" , f"{absentie[0]}" , "ABSENT" , "ABSENT" , "ABSENT" , f"{absentie2[0]}"])
        absentie.pop(0)
        absentie2.pop(0)

    wb.save(f"./marksheets/consise_marksheet.xlsx")
    flash("generated consise marksheet")
    return render_template("project1.html" , kjr="joel")




app.secret_key = b'_5#nbm2jk"F4Q8z\n\kjl\n]/'

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USERNAME'] = 'skyh2477@gmail.com'
app.config['MAIL_PASSWORD'] = 'IPL@J8143818'
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
mail = Mail(app)
@app.route("/lkj", methods=["POST", "GET"])
def mail1():
    if flask.request.method == "POST":
        with open("./storage/responses.csv" , "r") as f:
            x= csv.reader(f , delimiter=",")
            for y in x:
                if y[3]=="Name":
                    pass
                elif y[6]=="ANSWER":
                    pass
                else:
                    print(y[4] , "joel")
                    msg = Message('Quiz1 marks', sender='skyh2477@gmail.com', recipients=[f"{y[4]}"])
                    with app.open_resource(f"./marksheets/{y[6]}.xlsx") as fp:
                        msg.body=""
                        msg.attach(f"./marksheets/{y[6]}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fp.read())
                    print("kkk")
                    print(f"{y[4]}")
                    mail.send(msg)
                    print("jjj")
                    msg = Message('Quiz1 marks', sender='skyh2477@gmail.com', recipients=[f"{y[1]}"])
                    with app.open_resource(f"./marksheets/{y[6]}.xlsx") as fp:
                        msg.body=""
                        msg.attach(f"./marksheets/{y[6]}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fp.read())
                    print(f"{y[1]}")
                    mail.send(msg)
    return  render_template("project1.html"  , name=23 )




if __name__ == '__main__':
    app.run(debug=True)
