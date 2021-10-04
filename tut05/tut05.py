def generate_marksheet():
    import csv
    from openpyxl import Workbook
    from openpyxl import load_workbook
    import os
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill
    if(os.path.exists("output")) : 
        pass
    else :
            os.mkdir("output")

    with open(f"names-roll.csv" , "r") as f:
        reader = csv.reader(f  , delimiter=',')
        l=[]
        name=[]
        for r in reader:
            if r[0]== "Roll":
                continue
            l.append(r[0])
            name.append(r[1])
    with open(f"subjects_master.csv" , "r") as f:
        reader = csv.reader(f , delimiter=',')
        d={}
        for r in reader:
            if r[0]== 'subno':
                continue
            d[r[0]]=[r[1] , r[2] , r[3] ]
    with open(f"grades.csv" , 'r') as f:
        i=2
        rea = csv.reader(f)
        list111=[]
        for r in rea:
            if r[0] == 'Roll':
                continue
            if(os.path.exists(f"output\\{r[0]}+{r[1]}.csv")):
                    with open(f"output\\{r[0]}+{r[1]}.csv" , "a" , newline='') as f:
                        fie = [ 'Subject No' , 'Subject Name' , 'L-T-P' , 'Credit' , 'Subject Type' , 'Grade']
                        writer = csv.DictWriter(f , fieldnames=fie)
                        writer.writerow({ 'Subject No':r[2]  , 'Subject Name':d[r[2]][0] , 'L-T-P':d[r[2]][1] , 'Credit':r[3]  , 'Subject Type':r[5] , 'Grade':r[4]})
            else:
                list111.append(r[1])
                with open(f"output\\{r[0]}+{r[1]}.csv" , "w" , newline='') as f:
                        fie = [ 'Subject No' , 'Subject Name' , 'L-T-P' , 'Credit' , 'Subject Type' , 'Grade']
                        writer = csv.DictWriter(f , fieldnames=fie)
                        writer.writeheader()
                        writer.writerow({ 'Subject No':r[2]  , 'Subject Name':d[r[2]][0] , 'L-T-P':d[r[2]][1] ,'Credit':r[3] ,  'Subject Type':r[5] , 'Grade':r[4]})
    n =l.copy()
    m = l.copy()
    mnn= l.copy()
    while(len(l)>0):
        if( not os.path.exists(f"output\\{l[0]}.xlsx")):
            wb = Workbook()
            sheet= wb.active
            sheet.title="overall"
            wb.save(f"output\\{l[0]}.xlsx")
            l.pop(0)
    l=True
    dict={}

    while(len(n)>0 ):
            wb = load_workbook(f"output\\{n[0]}.xlsx")
            s=0
            while(True):
                if(len(list111)==0):
                    l= False
                    break
                if( not os.path.exists(f"output\\{n[0]}+{(list111[0])}.csv")):
                    dict[n[0]]=s
                    n.pop(0)
                    break
                else:
                    with open(f"output\\{n[0]}+{list111[0]}.csv" , "r" , newline='') as f:
                        wb.create_sheet(index=int(list111[0])  , title="Sem"+str(list111[0]))
                        sheet= wb["Sem"+str(list111[0])]
                        s+=1
                        roww = csv.reader(f , delimiter=',')
                        for item in roww:
                            if item[0]=='Subject No':
                                item = ['SI No']+item
                            else :
                                roww = sheet.max_row
                                item = [roww]+item
                            sheet.append(item)
                        if int(list111[0]) == 2:
                            for row in sheet.iter_rows(min_row=1, max_row=1):
                                for cell in row:
                                    cell.font = Font(bold=True)
                                    cell.fill = PatternFill(fgColor="77C3FD", fill_type = "solid")
                            row_count = sheet.max_row
                            for row in sheet.iter_rows(min_row=2,max_row=row_count):
                                for cell in row:
                                    cell.font = Font(bold=True)
                                    cell.fill = PatternFill(fgColor="B5DDFB", fill_type = "solid")
                        wb.save(f"output\\{n[0]}.xlsx")
                    os.remove(f"output\\{n[0]}+{list111[0]}.csv")
                    list111.pop(0)
            if l == False:
                dict[n[0]]=s
                break
    credits_taken={x:[]  for x in mnn}
    dictionary__={x:[] for x in mnn }
    semester_wise={x:[] for x  in mnn}
    roll_values={}
    d={'AA':10 , 'AB':9 , 'BB':8 ,'BC':7 , 'CC':6 , 'CD':5 , 'DD':4 , 'F':0 , 'I':0  , 'I*':0 , "F*":0 , 'DD*':0}
    while(len(m) >0):
        l11=[]
        l12=[]
        i=1
        wb = load_workbook(f"output\\{m[0]}.xlsx")
        xc=dict[m[0]]
        total=0
        while( xc >= i ):
            if m[0]== '0401ME11':
                if i==9:
                    i=i+1
            sheet = wb["Sem"+str(i)]
            for x in  range (2 , sheet.max_row+1) :
                l11.append(d[(sheet.cell(row=x, column=7).value).strip()])
                l12.append(int(sheet.cell(row=x, column=5).value))
            i+=1
            p=0
            q=0
            while len(l11) > 0 and len(l12)>0:
                p+=(l11[0]*l12[0])
                q+=l12[0]
                l11.pop(0)
                l12.pop(0)
            total +=q
            semester_wise[m[0]].append(total)
            dictionary__[m[0]].append(p/q)
            credits_taken[m[0]].append(q)
        m.pop(0)
    kl= dictionary__.copy()
    for x , y in dictionary__.items():
        i=0
        for v in y :
            dictionary__[x][i]= round(v , 2)
            i+=1
    dcv ={x : [] for x in mnn}
    bn= mnn.copy()
    while len(bn) > 0:
        kgf1 =0
        kgf2=0
        s=dict[bn[0]]
        i=0
        c=0
        while i < s:
            kgf1 +=kl[bn[0]][i]*credits_taken[bn[0]][i]
            kgf2+=credits_taken[bn[0]][i]
            c= kgf1/kgf2
            d11=round(c , 2)
            dcv[bn[0]].append(d11)
            i+=1
        bn.pop(0)
    while len(mnn) >0:
            wb =load_workbook(f"output\\{mnn[0]}.xlsx")
            s= dict[mnn[0]]
            sheet = wb["overall"]
            sheet.append(["Roll  No. " ,  mnn[0] ])
            sheet.append(["Name of The Student ", name[0] ])
            n = mnn[0] 
            sheet.append(["Discipline " ,  n[4:6]])
            if s==1:
                sheet.append(["Semester No" , 1])
                sheet.append(["Semester wise Credit taken" , credits_taken[mnn[0]][0]])
                sheet.append(["SPI"  , dictionary__[mnn[0]][0]])
                sheet.append(["Credits Taken" , semester_wise[mnn[0]][0]])
                sheet.append(["CPI"  , dcv[mnn[0]][0]  ])
            elif s==2:
                sheet.append(["Semester No" , 1 , 2])
                sheet.append(["Semester wise Credit taken" , credits_taken[mnn[0]][0] , credits_taken[mnn[0]][1]])
                sheet.append(["SPI"  , dictionary__[mnn[0]][0] , dictionary__[mnn[0]][1]])
                sheet.append(["Credits Taken" , semester_wise[mnn[0]][0] , semester_wise[mnn[0]][1]])
                sheet.append(["CPI"  , dcv[mnn[0]][0]  , dcv[mnn[0]][1] ])
            elif s==3:
                sheet.append(["Semester No" , 1 , 2 , 3])
                sheet.append(["Semester wise Credit taken" , credits_taken[mnn[0]][0] , credits_taken[mnn[0]][1], credits_taken[mnn[0]][2]])
                sheet.append(["SPI"  , dictionary__[mnn[0]][0] ,dictionary__[mnn[0]][1] , dictionary__[mnn[0]][2]])
                sheet.append(["Credits Taken" , semester_wise[mnn[0]][0] , semester_wise[mnn[0]][1] , semester_wise[mnn[0]][2]])
                sheet.append(["CPI"  , dcv[mnn[0]][0]  , dcv[mnn[0]][1] , dcv[mnn[0]][2]])
            elif s==4:
                sheet.append(["Semester No" , 1 , 2, 3 , 4])
                sheet.append(["Semester wise Credit taken" , credits_taken[mnn[0]][0] , credits_taken[mnn[0]][1] ,credits_taken[mnn[0]][2] , credits_taken[mnn[0]][3]])
                sheet.append(["SPI"  , dictionary__[mnn[0]][0] , dictionary__[mnn[0]][1] , dictionary__[mnn[0]][2] , dictionary__[mnn[0]][3]])
                sheet.append(["Credits Taken" , semester_wise[mnn[0]][0] , semester_wise[mnn[0]][1] , semester_wise[mnn[0]][2] , semester_wise[mnn[0]][3]])
                sheet.append(["CPI"  , dcv[mnn[0]][0]  , dcv[mnn[0]][1] , dcv[mnn[0]][2] ,  dcv[mnn[0]][3]  ])
            elif s==5:
                sheet.append(["Semester No" , 1 , 2,3,4 ,5])
                sheet.append(["Semester wise Credit taken" , credits_taken[mnn[0]][0] , credits_taken[mnn[0]][1] , credits_taken[mnn[0]][2] , credits_taken[mnn[0]][3] , credits_taken[mnn[0]][4]])
                sheet.append(["SPI"  , dictionary__[mnn[0]][0] , dictionary__[mnn[0]][1] , dictionary__[mnn[0]][2] , dictionary__[mnn[0]][3] , dictionary__[mnn[0]][4]])
                sheet.append(["Credits Taken" , semester_wise[mnn[0]][0] , semester_wise[mnn[0]][1] , semester_wise[mnn[0]][2] , semester_wise[mnn[0]][3], semester_wise[mnn[0]][4]])
                sheet.append(["CPI"  , dcv[mnn[0]][0]  , dcv[mnn[0]][1] , dcv[mnn[0]][2] ,  dcv[mnn[0]][3] , dcv[mnn[0]][4] ])
            elif s==6:
                sheet.append(["Semester No" , 1, 2 , 3, 4 , 5 , 6])
                sheet.append(["Semester wise Credit taken" , credits_taken[mnn[0]][0] ,credits_taken[mnn[0]][1], credits_taken[mnn[0]][2] , credits_taken[mnn[0]][3] , credits_taken[mnn[0]][4] , credits_taken[mnn[0]][5]])
                sheet.append(["SPI"  , dictionary__[mnn[0]][0] , dictionary__[mnn[0]][1] , dictionary__[mnn[0]][2] , dictionary__[mnn[0]][3] , dictionary__[mnn[0]][4] , dictionary__[mnn[0]][5]])
                sheet.append(["Credits Taken" , semester_wise[mnn[0]][0] , semester_wise[mnn[0]][1] , semester_wise[mnn[0]][2] , semester_wise[mnn[0]][3], semester_wise[mnn[0]][4] , semester_wise[mnn[0]][5]])
                sheet.append(["CPI"  , dcv[mnn[0]][0]  , dcv[mnn[0]][1] , dcv[mnn[0]][2] ,  dcv[mnn[0]][3] , dcv[mnn[0]][4] ,dcv[mnn[0]][5] ])
            elif s==7:
                sheet.append(["Semester No" , 1 , 2 , 3 , 4, 5, 6, 7])
                sheet.append(["Semester wise Credit taken" , credits_taken[mnn[0]][0] ,credits_taken[mnn[0]][1], credits_taken[mnn[0]][2] , credits_taken[mnn[0]][3] , credits_taken[mnn[0]][4] , credits_taken[mnn[0]][5] ,  credits_taken[mnn[0]][6]])
                sheet.append(["SPI"  , dictionary__[mnn[0]][0] , dictionary__[mnn[0]][1] , dictionary__[mnn[0]][2] , dictionary__[mnn[0]][3] , dictionary__[mnn[0]][4] , dictionary__[mnn[0]][5] , dictionary__[mnn[0]][6]])
                sheet.append(["Credits Taken" , semester_wise[mnn[0]][0] , semester_wise[mnn[0]][1] , semester_wise[mnn[0]][2] , semester_wise[mnn[0]][3], semester_wise[mnn[0]][4] , semester_wise[mnn[0]][5] , semester_wise[mnn[0]][6]])
                sheet.append(["CPI"  , dcv[mnn[0]][0]  , dcv[mnn[0]][1] , dcv[mnn[0]][2] ,  dcv[mnn[0]][3] , dcv[mnn[0]][4] ,dcv[mnn[0]][5],dcv[mnn[0]][6]  ])
            elif mnn[0]=='0401ME11':
                sheet.append(["Semester No" , 1 , 2 , 3,4,5 ,6,7,8,10 ])
                sheet.append(["Semester wise Credit taken" , credits_taken[mnn[0]][0] ,credits_taken[mnn[0]][1], credits_taken[mnn[0]][2] , credits_taken[mnn[0]][3] , credits_taken[mnn[0]][4] , credits_taken[mnn[0]][5] , credits_taken[mnn[0]][6]  , credits_taken[mnn[0]][7]])
                sheet.append(["SPI"  , dictionary__[mnn[0]][0] , dictionary__[mnn[0]][1] , dictionary__[mnn[0]][2] , dictionary__[mnn[0]][3] , dictionary__[mnn[0]][4] , dictionary__[mnn[0]][5] , dictionary__[mnn[0]][6] , dictionary__[mnn[0]][7]])       
                sheet.append(["Credits Taken" , semester_wise[mnn[0]][0] , semester_wise[mnn[0]][1] , semester_wise[mnn[0]][2] , semester_wise[mnn[0]][3], semester_wise[mnn[0]][4] , semester_wise[mnn[0]][5] , semester_wise[mnn[0]][6] , semester_wise[mnn[0]][7]])
                sheet.append(["CPI"  , dcv[mnn[0]][0]  , dcv[mnn[0]][1] , dcv[mnn[0]][2] ,  dcv[mnn[0]][3] , dcv[mnn[0]][4] ,dcv[mnn[0]][5] , dcv[mnn[0]][6] , dcv[mnn[0]][7]])
            elif s==8: 
                sheet.append(["Semester No" , 1 , 2 , 3 , 4 , 5 , 6  , 7 , 8 ] )
                sheet.append(["Semester wise Credit taken" , credits_taken[mnn[0]][0] ,credits_taken[mnn[0]][1], credits_taken[mnn[0]][2] , credits_taken[mnn[0]][3] , credits_taken[mnn[0]][4] , credits_taken[mnn[0]][5] , credits_taken[mnn[0]][6]  , credits_taken[mnn[0]][7]])
                sheet.append(["SPI"  , dictionary__[mnn[0]][0] , dictionary__[mnn[0]][1] , dictionary__[mnn[0]][2] , dictionary__[mnn[0]][3] , dictionary__[mnn[0]][4] , dictionary__[mnn[0]][5] , dictionary__[mnn[0]][6] , dictionary__[mnn[0]][7]])        
                sheet.append(["Credits Taken" , semester_wise[mnn[0]][0] , semester_wise[mnn[0]][1] , semester_wise[mnn[0]][2] , semester_wise[mnn[0]][3], semester_wise[mnn[0]][4] , semester_wise[mnn[0]][5] , semester_wise[mnn[0]][6] , semester_wise[mnn[0]][7]])
                sheet.append(["CPI"  , dcv[mnn[0]][0]  , dcv[mnn[0]][1] , dcv[mnn[0]][2] ,  dcv[mnn[0]][3] , dcv[mnn[0]][4] ,dcv[mnn[0]][5] , dcv[mnn[0]][6] , dcv[mnn[0]][7]])
            else:
                sheet.append(["Semester No" , 1 , 2 , 3 , 4 , 5 , 6 , 7, 8, 9])
                sheet.append(["Semester wise Credit taken" , credits_taken[mnn[0]][0] ,credits_taken[mnn[0]][1], credits_taken[mnn[0]][2] , credits_taken[mnn[0]][3] , credits_taken[mnn[0]][4] , credits_taken[mnn[0]][5] , credits_taken[mnn[0]][6] , credits_taken[mnn[0]][7] , credits_taken[mnn[0]][8]])
                sheet.append(["SPI"  , dictionary__[mnn[0]][0] , dictionary__[mnn[0]][1] , dictionary__[mnn[0]][2] , dictionary__[mnn[0]][3] , dictionary__[mnn[0]][4] , dictionary__[mnn[0]][5] , dictionary__[mnn[0]][6] , dictionary__[mnn[0]][7] , dictionary__[mnn[0][8]]])
                sheet.append(["Credits Taken" , semester_wise[mnn[0]][0] , semester_wise[mnn[0]][1] , semester_wise[mnn[0]][2] , semester_wise[mnn[0]][3], semester_wise[mnn[0]][4] , semester_wise[mnn[0]][5] , semester_wise[mnn[0]][6] , semester_wise[mnn[0]][7] , semester_wise[mnn[0]][8]])
                sheet.append(["CPI"  , dcv[mnn[0]][0]  , dcv[mnn[0]][1] , dcv[mnn[0]][2] ,  dcv[mnn[0]][3] , dcv[mnn[0]][4] ,dcv[mnn[0]][5] , dcv[mnn[0]][6] , dcv[mnn[0]][7] ,dcv[mnn[0]][8]   ])
            wb.save(f"output\\{mnn[0]}.xlsx")
            mnn.pop(0)
            name.pop(0)
    return
generate_marksheet()