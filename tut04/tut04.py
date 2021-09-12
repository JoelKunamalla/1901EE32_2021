 #  KUNAMALLA JOEL RATHNAM 1901EE32
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
def output_by_subject() :
    import os
    list11=[]
    #os.mkdir("output_by_subject")
    if(os.path.exists("output_by_subject")) : 
        pass
    else :
            os.mkdir("output_by_subject")
    with open("regtable_old.csv", "r") as f:
        roww = csv.reader(f , delimiter=',')
        row=0
        for r in roww:
            if r[0][0]=='r':
                continue
            list11.append(r[3])
            if(os.path.exists(f"output_by_subject\\{r[3]}.csv")):
                with open(f"output_by_subject\\{r[3]}.csv" , "a" , newline='') as f:
                     f1 = ['rollno' , 'register_sem' , 'subno' , 'sub_type']
                     writer = csv.DictWriter(f , fieldnames=f1)
                     writer.writerow({'rollno':r[0] , 'register_sem':r[1] , 'subno':r[3] , 'sub_type':r[8]})
            else:
                with open(f"output_by_subject\\{r[3]}.csv" , "w" , newline='') as f:
                    fie = ['rollno' , 'register_sem' , 'subno' , 'sub_type']
                    writer = csv.DictWriter(f , fieldnames=fie)
                    writer.writeheader()
                    writer.writerow({'rollno':r[0] , 'register_sem':r[1]  , 'subno':r[3] , 'sub_type':r[8]})

    while len(list11) > 0 :
        if( not os.path.exists(f"output_by_subject\\{list11[0]}.xlsx")):
            wb = Workbook()
            sheet=  wb.active
            with open(f"output_by_subject\\{list11[0]}.csv" , "r" , newline='') as f:
                roww = csv.reader(f , delimiter=',')
                for item in roww:
                    sheet.append(item)
                wb.save(f"output_by_subject\\{list11[0]}.xlsx")
            os.remove(f"output_by_subject\\{list11[0]}.csv")
        list11.pop(0)
    return 



def output_individual_roll() :
    import os
    list11=[]
    #os.mkdir("output_by_subject")
    if(os.path.exists("output_individual_roll")) : 
        pass
    else :
            os.mkdir("output_individual_roll")
    with open("regtable_old.csv", "r") as f:
        roww = csv.reader(f , delimiter=',')
        row=0
        for r in roww:
            if r[0][0]=='r':
                continue
            list11.append(r[0])
            if(os.path.exists(f"output_individual_roll\\{r[0]}.csv")):
                with open(f"output_individual_roll\\{r[0]}.csv" , "a" , newline='') as f:
                     f1 = ['rollno' , 'register_sem' , 'subno' , 'sub_type']
                     writer = csv.DictWriter(f , fieldnames=f1)
                     writer.writerow({'rollno':r[0] , 'register_sem':r[1] , 'subno':r[3] , 'sub_type':r[8]})
            else:
                with open(f"output_individual_roll\\{r[0]}.csv" , "w" , newline='') as f:
                    fie = ['rollno' , 'register_sem' , 'subno' , 'sub_type']
                    writer = csv.DictWriter(f , fieldnames=fie)
                    writer.writeheader()
                    writer.writerow({'rollno':r[0] , 'register_sem':r[1]  , 'subno':r[3] , 'sub_type':r[8]})

    while len(list11) > 0 :
        if( not os.path.exists(f"output_individual_roll\\{list11[0]}.xlsx")):
            wb = Workbook()
            sheet=  wb.active
            with open(f"output_individual_roll\\{list11[0]}.csv" , "r" , newline='') as f:
                roww = csv.reader(f , delimiter=',')
                for item in roww:
                    sheet.append(item)
                wb.save(f"output_individual_roll\\{list11[0]}.xlsx")
            os.remove(f"output_individual_roll\\{list11[0]}.csv")
        list11.pop(0)

    return 
output_by_subject()
output_individual_roll()